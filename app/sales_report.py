from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.config import REPORT_DIR, SHOP_ID
from datetime import datetime
from app.db.database import get_sales_data, get_sale_items_data,get_daily_metrics, get_top_products_by_revenue, get_top_products_by_weight, get_top_products_by_pieces
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.protection import SheetProtection

# ============================================================
# DESIGN TOKENS — define once, reuse everywhere
# ============================================================
# Colors
COLOR_PRIMARY = '1F3864'       # Dark blue — title bar
COLOR_ACCENT = '374151'        # Dark grey with slight blue tint — section headers
COLOR_BG_LIGHT = 'F4F6FA'      # Very light blue — subtle backgrounds
COLOR_WHITE = 'FFFFFF'         # White text
COLOR_TEXT = '262626'          # Near-black body text
COLOR_MUTED = '595959'         # Grey for subtitles/footer
COLOR_BORDER = 'BFBFBF'        # Light grey borders

# Fonts
FONT_NAME = 'Calibri'
FONT_TITLE_SIZE = 16
FONT_SECTION_SIZE = 12
FONT_BODY_SIZE = 10
FONT_FOOTER_SIZE = 9

# Number formats
FMT_CURRENCY = '"₹"#,##0.00'
FMT_CURRENCY_INT = '"₹"#,##0'
FMT_INT = '#,##0'
FMT_DATE = 'yyyy-mm-dd'
FMT_TIME = 'hh:mm'

# Border style (thin grey)
THIN_BORDER = Border(
    left=Side(style='thin', color="BFBFBF"),
    right=Side(style='thin', color="BFBFBF"),
    top=Side(style='thin', color="BFBFBF"),
    bottom=Side(style='thin', color="BFBFBF"),
)

# ============================================================
# HELPER FUNCTIONS
# ============================================================


def style_title(cell, size=18):
    """Title cell — large, bold, dark blue."""
    cell.font = Font(name=FONT_NAME, size=size, bold=True, color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def style_subtitle(cell):
    """Subtitle — smaller, muted grey."""
    cell.font = Font(name=FONT_NAME, size=10, color=COLOR_MUTED, italic=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')


def style_section_header(cell):
    """Section header — bold white on accent background."""
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill('solid', start_color=COLOR_ACCENT)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

def style_table_header(cell):
    """Table column header — bold white on primary background."""
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_WHITE)
    cell.fill = PatternFill('solid', start_color=COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

def style_data_cell(cell, bg_color=None):
    """Regular data cell — light border, optional background."""
    cell.font = Font(name=FONT_NAME, size=10)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical='center')
    if bg_color:
        cell.fill = PatternFill('solid', start_color=bg_color)

def set_up_sheet_view(ws,filter_row: int = None, filter_col_range: tuple[int, int] = None,freeze_top_rows: str = None):
    # Freeze top rows (title + headers) so user can scroll long lists
    if freeze_top_rows:
        ws.freeze_panes = freeze_top_rows
    
    # Hide gridlines for cleaner look
    ws.sheet_view.showGridLines = False

    
    # protect sheet
    ws.protection.enabled = True
    # Protect sheet but explicitly allow autofilter and sort
    ws.protection = SheetProtection(
        sheet=True,
        autoFilter=False,   # False = allow autofilter
        sort=False,         # False = allow sort
        selectLockedCells=False,
        selectUnlockedCells=False,
    )   

    # Add filters to the headers (e.g., Row 2 covers columns A to I)
    if (filter_row) and (filter_col_range):
        col1,col2 = filter_col_range
        ws.auto_filter.ref = f"{get_column_letter(col1)}{filter_row}:{get_column_letter(col2)}{filter_row}" 
        ws.auto_filter.enable = True

    # 2. Lock all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=(cell.row != filter_row))

def build_sales_list_sheet(wb, timestamp):
  

    # Query database
    sales = get_sales_data()

    # Define headers - user-friendly
    headers = [
            "Bill #", "Date", "Time",  "Payment Mode", "Total",
            "Items", "Type(Sale/Refund)", "Refund For(Bill #)", "Voided"
        ] 
    
    # Layout constants — self-documenting
    TITLE_ROW = 1
    HEADER_ROW = 3
    DATA_START_ROW = 4

    # Build workbook
    # wb = Workbook()
    # ws = wb.active
    ws = wb.create_sheet("Sales List")

    # Column widths
    widths = {'A': 14, 'B': 14, 'C': 12, 'D': 16, 'E': 14, 'F': 14, 'G': 18, 'H': 18, 'I': 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Title (row 1)
    last_col = get_column_letter(len(headers))
    title_cell = ws.cell(row =TITLE_ROW, column = 1, value = f"Sales List — {timestamp}")
    style_title(title_cell, size=14)
    ws.merge_cells(f"A{TITLE_ROW}:{last_col}{TITLE_ROW}")
    ws.row_dimensions[TITLE_ROW].height = 24

    # Headers (row 3)
    for col_index, header in enumerate(headers, start = 1):
        header_cell = ws.cell(row=HEADER_ROW, column=col_index, value=header)
        style_table_header(header_cell)
    ws.row_dimensions[HEADER_ROW].height = 28
    
    # Data rows
    if not sales:
        ws.cell(row=DATA_START_ROW, column=1, value="No sales recorded")
        
        return

    for row_offset, sale in enumerate(sales):
        row = DATA_START_ROW + row_offset
        
        ws.cell(row=row, column=1, value=sale['bill_number'])
        ws.cell(row=row, column=2, value=sale['timestamp'][:10])  # date
        ws.cell(row=row, column=3, value=sale['timestamp'][11:16])  # time
        ws.cell(row=row, column=4, value=sale['payment_mode'])
        ws.cell(row=row, column=5, value=sale['total_price'])
        ws.cell(row=row, column=6, value=sale['items_count'])
        ws.cell(row=row, column=7, value=sale['transaction_type'])
        ws.cell(row=row, column=8, value=sale['refund_for_bill'])
        ws.cell(row=row, column=9, value='Yes' if sale['is_void'] else 'No')

        # Apply styling to all cells in this row
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            style_data_cell(cell)
        # Specific formats per column
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=5).number_format = FMT_CURRENCY
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=9).alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row].height = 18
    # Totals row
    data_end_row = DATA_START_ROW + len(sales) - 1
    totals_row = data_end_row + 2

    # Label spanning columns 1-3
    ws.cell(row=totals_row, column=1, value="Net Total")
    ws.cell(row=totals_row, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=totals_row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(
        start_row=totals_row, start_column=1,
        end_row=totals_row, end_column=3
    )
    
    # SUMIF formula — uses Yes/No instead of 0/1 because we changed display
    void_col = 9   # column 9 — "Voided" column
    total_col = 5  # column 4 — "Total" column
    
    formula = (
        f'=SUMIF({get_column_letter(void_col)}{DATA_START_ROW}:{get_column_letter(void_col)}{data_end_row},"No",'
        f'{get_column_letter(total_col)}{DATA_START_ROW}:{get_column_letter(total_col)}{data_end_row})'
    )
    ws.cell(row=totals_row, column=total_col, value=formula)
    ws.cell(row=totals_row, column=total_col).font = Font(name=FONT_NAME, size=11, bold=True, color="000000")
    ws.cell(row=totals_row, column=total_col).number_format = FMT_CURRENCY_INT
    ws.cell(row=totals_row, column=total_col).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=totals_row, column=total_col).fill = PatternFill('solid', start_color=COLOR_BG_LIGHT)
    ws.cell(row=totals_row, column=total_col).border = Border(top=Side(style='medium', color=COLOR_PRIMARY))

    ws.row_dimensions[totals_row].height = 24

    set_up_sheet_view(ws,filter_row=HEADER_ROW, filter_col_range=(1,9),freeze_top_rows='A4')

def build_items_detail_sheet(wb, timestamp):

    # Query database
    sale_items = get_sale_items_data()

    # Define headers - user-friendly
    headers = [
            "Bill #", "is_void", "Product", "Local Name", "Quantity", "Unit", "Packets", "Line Total"
        ] 
    
    # Layout constants — self-documenting
    TITLE_ROW = 1
    HEADER_ROW = 3
    DATA_START_ROW = 4

    ws = wb.create_sheet("Items Detail")

    # Column widths
    widths = {'A': 14, 'B': 10, 'C': 20, 'D': 20, 'E': 14, 'F': 14, 'G': 14, 'H': 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Title (row 1)
    last_col = get_column_letter(len(headers))
    title_cell = ws.cell(row =TITLE_ROW, column = 1, value = f"Items Detail — {timestamp}")
    style_title(title_cell, size=14)
    ws.merge_cells(f"A{TITLE_ROW}:{last_col}{TITLE_ROW}")
    ws.row_dimensions[TITLE_ROW].height = 24

    # Headers (row 3)
    for col_index, header in enumerate(headers, start = 1):
        header_cell = ws.cell(row=HEADER_ROW, column=col_index, value=header)
        style_table_header(header_cell)
    ws.row_dimensions[HEADER_ROW].height = 28
    ws.auto_filter.ref = f"A{HEADER_ROW}:H{HEADER_ROW}" # Add filters to the headers
    
    # Data rows
    if not sale_items:
        ws.cell(row=DATA_START_ROW, column=1, value="No sale items recorded")
        
        return
    
    for row_offset, sale_item in enumerate(sale_items):
        row = DATA_START_ROW + row_offset
        
        ws.cell(row=row, column=1, value=sale_item['bill_number'])
        ws.cell(row=row, column=2, value='Yes' if sale_item['is_void'] else 'No')
        ws.cell(row=row, column=3, value=sale_item['name']) 
        ws.cell(row=row, column=4, value=sale_item['local_name']) 
        ws.cell(row=row, column=5, value=sale_item['quantity'])
        ws.cell(row=row, column=6, value=sale_item['unit'])
        ws.cell(row=row, column=7, value=sale_item['packets'])
        ws.cell(row=row, column=8, value=sale_item['line_total'])

        # Apply styling to all cells in this row
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            style_data_cell(cell)
        # Specific formats per column
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=5).number_format = FMT_CURRENCY
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=8).number_format = FMT_CURRENCY
        

        ws.row_dimensions[row].height = 18

    # Totals row
    data_end_row = DATA_START_ROW + len(sale_items) - 1
    totals_row = data_end_row + 2

    # Label spanning columns 1-3
    ws.cell(row=totals_row, column=1, value="Grand Total")
    ws.cell(row=totals_row, column=1).font = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_PRIMARY)
    ws.cell(row=totals_row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells(
        start_row=totals_row, start_column=1,
        end_row=totals_row, end_column=7
    )
    
    # SUM formula
    total_col = 8  # column 4 — "Total" column
    
    formula = (
        f"=SUM({get_column_letter(total_col)}{DATA_START_ROW}:{get_column_letter(total_col)}{data_end_row})")
    
    ws.cell(row=totals_row, column=total_col, value=formula)
    ws.cell(row=totals_row, column=total_col).font = Font(name=FONT_NAME, size=11, bold=True, color="000000")
    ws.cell(row=totals_row, column=total_col).number_format = FMT_CURRENCY_INT
    ws.cell(row=totals_row, column=total_col).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=totals_row, column=total_col).fill = PatternFill('solid', start_color=COLOR_BG_LIGHT)
    ws.cell(row=totals_row, column=total_col).border = Border(top=Side(style='medium', color=COLOR_PRIMARY))

    ws.row_dimensions[totals_row].height = 24

    set_up_sheet_view(ws,filter_row=HEADER_ROW, filter_col_range=(1,8),freeze_top_rows='A4')

def build_daily_charts(ws,data_ws,chart_1_cell_placement, chart_2_cell_placement):
    """Build the daily time-series charts on the Summary sheet."""
    daily = get_daily_metrics(days=30)
    
    if not daily:
        ws.cell(row=1, column=1, value="No sales in the last 30 days")
        return
    
    # Layout constants
    HEADER_ROW = 1
    DATA_START_ROW = 2
    
    # Write headers
    headers = list(daily[0].keys())
    data_ws.append(headers)
    
    # Write data
    for row in daily:
        data_ws.append(list(row.values()))
    
    last_data_row = HEADER_ROW + len(daily)
    
    # Categories reference — dates in column A
    categories = Reference(data_ws, min_col=1, min_row=DATA_START_ROW, max_row=last_data_row)
    
    # Chart 1: Daily Revenue
    revenue_chart = BarChart()
    revenue_chart.title = "Daily Revenue (Last 30 Days)"
    revenue_chart.y_axis.title = "Revenue (₹)"
    revenue_chart.x_axis.title = "Date"
    revenue_chart.legend = None
    
    revenue_data = Reference(data_ws, min_col=2, min_row=HEADER_ROW, max_row=last_data_row)
    revenue_chart.add_data(revenue_data, titles_from_data=True)
    revenue_chart.set_categories(categories)
    
    # Chart 2: Cash vs GPay
    payment_chart = BarChart()
    payment_chart.title = "Cash vs GPay (Last 30 Days)"
    payment_chart.y_axis.title = "Amount (₹)"
    payment_chart.x_axis.title = "Date"
    
    payment_data = Reference(data_ws, min_col=4, min_row=HEADER_ROW, max_col=5, max_row=last_data_row)
    payment_chart.add_data(payment_data, titles_from_data=True)
    payment_chart.set_categories(categories)
    
    # Position charts to the RIGHT of data (columns 1-7 have data)
    ws.add_chart(revenue_chart, chart_1_cell_placement)
    ws.add_chart(payment_chart, chart_2_cell_placement)

    return last_data_row

def build_top_products_chart(
      ws, 
      data_ws, 
      start_row: int, 
      top_product_func, 
      limit: int, 
      period: str, 
      chart_title: str, 
      chart_y_title: str, 
      cell_placement: str
    ) -> int:
   
    """Build the top products charts on the Summary sheet."""
    top_products = top_product_func(limit=limit, period=period)

    header_row = start_row + 1
    first_data_row = header_row + 1
    last_data_row = first_data_row + len(top_products) - 1
    

    if not top_products:
        ws.cell(row=header_row, column=1, value=f"No data for {chart_title}")
        return header_row
    
    # Write headers
    headers = list(top_products[0].keys())
    data_ws.append(headers)

    # Write data
    for row in top_products:
        data_ws.append(list(row.values()))
    
    

    # Categories reference — dates in column A
    categories = Reference(data_ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
    
    # Chart 1
    top_products_chart = BarChart()
    top_products_chart.title = chart_title
    top_products_chart.y_axis.title = chart_y_title
    top_products_chart.legend = None
    
    top_products_data = Reference(data_ws, min_col=2, min_row=header_row, max_row=last_data_row)
    
    top_products_chart.add_data(top_products_data, titles_from_data=True)
    top_products_chart.set_categories(categories)

    ws.add_chart(top_products_chart, cell_placement)
    return last_data_row

def build_kpi_box(ws, top_left_cell, kpi_height, label, value, context, color, number_format=None):
    """Draw a KPI box with label, value, and context text."""
    row, col = top_left_cell

    FONT_NAME = 'Calibri'
    TEXT_COLOR = '1F3864'      # Dark blue for value
    LABEL_TEXT_COLOR = 'FFFFFF' # White for header text on colored bg
    CONTEXT_COLOR = '595959'    # Muted grey for subtitle
    BORDER_COLOR = 'BFBFBF'     # Light grey border

    thin_border = Side(style='thin', color=BORDER_COLOR)
    box_border = Border(
        left=thin_border,
        right=thin_border,
        top=thin_border,
        bottom=thin_border,
    )
    
    # Label cell (colored header)
    label_cell = ws.cell(row=row, column=col, value=label.upper())

    label_cell.fill = PatternFill(
        fill_type='solid',
        start_color=color,
    )
    label_cell.font = Font(
        name=FONT_NAME,
        size=10,
        bold=True,
        color=LABEL_TEXT_COLOR,
    )
    label_cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
    )
    label_cell.border = box_border

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
    ws.row_dimensions[row].height = 22
    
    # Value cell (large number)
    value_cell = ws.cell(row=row+1, column=col, value=value)
    value_cell.font = Font(
        name=FONT_NAME,
        size=20,
        bold=True,
        color=TEXT_COLOR,
    )
    value_cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
    )
    if number_format is not None:
        value_cell.number_format = number_format

    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+3, end_column=col+3)
    
    # Context cell (subtitle)
    context_cell = ws.cell(row=kpi_height, column=col, value=context)
    context_cell.font = Font(
        name=FONT_NAME,
        size=9,
        italic=True,
        color=CONTEXT_COLOR,
    )
    context_cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
    )
    ws.merge_cells(start_row=kpi_height, start_column=col, end_row=kpi_height, end_column=col+3)
    ws.row_dimensions[kpi_height].height = 18
    
    # ==================================================
    # SIDE BORDERS (to close the "card" appearance)
    # ==================================================
    # Add left/right/bottom borders to intermediate cells to complete the card look.
    for r in range(row+1, kpi_height+1):
       for c in range(col,col+4):
        cell = ws.cell(row=r, column=c)
        existing = cell.border
        cell.border = Border(
            left=thin_border if c == col else existing.left,
            right=thin_border if c == (col+3) else existing.right,
            bottom=thin_border if r == (kpi_height) else existing.bottom,
        )
    kpi_column_next_box = col+5 # 5 columns spacing (4 cols wide + 1 gap)
    return kpi_column_next_box
        
       

def build_summary_sheet(wb, timestamp, data_sheet):
    ws = wb.create_sheet("Summary")
    # rows and colimns reference
    space_row = 2
    title_row = 1
    left_start_col = 2
    last_col = 20
    sub_title_row = title_row+1

    kpi_row = sub_title_row + space_row
    kpi_col_box_1 = left_start_col
    kpi_height = kpi_row+4

    daily_trend_header_row = kpi_height + space_row
    daily_chart_row = daily_trend_header_row + space_row
    side_chart_col = 12

    chart_height = 14
    performance_header_row = daily_chart_row + chart_height + space_row
    performance_chart_row = performance_header_row + space_row

    #Title block
    title_cell = ws.cell(row=title_row, column=left_start_col, value=f"HAPPYMAN SWEETS — Branch 1 | {timestamp}")
    style_title(title_cell)
    ws.merge_cells(start_row=title_row, start_column=left_start_col, end_row=title_row, end_column=last_col)
    ws.row_dimensions[title_row].height = 28

    sub_title_cell = ws.cell(row=sub_title_row, column=left_start_col, value=f"Daily Sales Report  |  Shop ID: {SHOP_ID}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    style_subtitle(sub_title_cell)
    ws.merge_cells(start_row=sub_title_row, start_column=left_start_col, end_row=sub_title_row, end_column=last_col)
    
    

    # ---------- KPI METRICS section ----------
    today = get_daily_metrics(days=1)[0]
    yesterday = get_daily_metrics(days=2)[0]
    # Box 1: Today's Revenue
    kpi_column_next_box = build_kpi_box(ws, top_left_cell=(kpi_row, kpi_col_box_1), 
                kpi_height = kpi_height,
                label="TODAY'S REVENUE",
                value=today['total_revenue'],
                context=f"yesterday: ₹{yesterday['total_revenue']:,.0f}",
                color='1F3864',
                number_format=FMT_CURRENCY_INT) # Dark blue
    # Box 2: Today's Sale Count
    kpi_column_next_box = build_kpi_box(ws, top_left_cell=(kpi_row, kpi_column_next_box), 
                kpi_height = kpi_height,
                label="TODAY'S SALES",
                value=today['sale_count'],
                context=f"yesterday: {yesterday['sale_count']}",
                color='2E7D32',
                number_format=FMT_INT)
    
    # Box 3: Today's Cash vs Gpay
    total = today['total_revenue']
    if total > 0:
        cash_pct = today['cash_total'] / total * 100
        value = f"{cash_pct:.0f}% cash / {100-cash_pct:.0f}% gpay"
        context = f"₹{today['cash_total']:,.0f} | ₹{today['gpay_total']:,.0f}"
    else:
        value = "—"  # em-dash indicating "no data"
        context = "No sales today yet"
    kpi_column_next_box = build_kpi_box(ws, top_left_cell=(kpi_row, kpi_column_next_box), 
                kpi_height = kpi_height,
                label="CASH vs GPAY (TODAY)",
                value=value,
                context=context,
                color='78350F',
                number_format=None) 
    
    # Box 4: Last 7 Days Revenue
    fourteen_days_data = get_daily_metrics(days=14)

    recent_revenue = sum(day['total_revenue'] for day in fourteen_days_data[-7:])
    previous_revenue = sum(day['total_revenue'] for day in fourteen_days_data[:7])

    if recent_revenue > 0:
        value = f"₹{recent_revenue:,.0f}"
        
        if previous_revenue > 0:
            change_pct = ((recent_revenue - previous_revenue) / previous_revenue) * 100
            if change_pct > 0:
                context = f"↑ {change_pct:.0f}% vs previous 7 days"
            elif change_pct < 0:
                context = f"↓ {abs(change_pct):.0f}% vs previous 7 days"
            else:
                context = "same as previous 7 days"
        else:
            context = "no sales in previous 7 days"
    else:
        value = "—"
        context = "no sales in last 7 days"

    kpi_column_next_box = build_kpi_box(
        ws,
        top_left_cell=(kpi_row, kpi_column_next_box),
        kpi_height=kpi_height,
        label="LAST 7 DAYS",
        value=value,
        context=context,
        color='0F766E',
        number_format=None,
    ) 

    #Section header: "DAILY TRENDS"
    daily_trend_cell = ws.cell(row=daily_trend_header_row, column=left_start_col, value="DAILY TRENDS")
    style_section_header(daily_trend_cell)
    ws.merge_cells(start_row=daily_trend_header_row, start_column=left_start_col, end_row=daily_trend_header_row, end_column=last_col)
    ws.row_dimensions[daily_trend_header_row].height = 22
    # Chart
    current_row  = build_daily_charts(ws, data_sheet,
                    chart_1_cell_placement=f"{get_column_letter(left_start_col)}{daily_chart_row}", 
                    chart_2_cell_placement=f"{get_column_letter(side_chart_col)}{daily_chart_row}")

    #Section header: "PRODUCT PERFORMANCE"
    performance_cell = ws.cell(row=performance_header_row, column=left_start_col, value="PRODUCT PERFORMANCE")
    style_section_header(performance_cell)
    ws.merge_cells(start_row=performance_header_row, start_column=left_start_col, end_row=performance_header_row, end_column=last_col)
    ws.row_dimensions[performance_header_row].height = 22
    # Chart
    current_row  = build_top_products_chart(ws, data_sheet, current_row ,get_top_products_by_weight,7, "last_30_days","Top Products by Weight (Last 30 Days)","Weight (kg)",  f"{get_column_letter(left_start_col)}{performance_chart_row}")
    current_row  = build_top_products_chart(ws, data_sheet, current_row ,get_top_products_by_pieces,7, "last_30_days","Top Products by Pieces (Last 30 Days)","Pieces",  f"{get_column_letter(side_chart_col)}{performance_chart_row}")

    
    current_row  = build_top_products_chart(ws, data_sheet, current_row ,get_top_products_by_revenue,7, "last_30_days","Top Products by Revenue (Last 30 Days)","Revenue (₹)",  f"{get_column_letter(left_start_col)}{performance_chart_row + chart_height + space_row}")
    # Turn off worksheet gridlines
    ws.sheet_view.showGridLines = False

    set_up_sheet_view(ws,freeze_top_rows='A2')

  

def main():
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"HappyMan_{SHOP_ID}_{timestamp}.xlsx"
    report_path = REPORT_DIR / filename

    wb = Workbook()
    wb.remove(wb.active)
    
    # Create hidden data sheet first (will be moved to end)
    data_sheet = wb.create_sheet("data_sheet")
    data_sheet.sheet_state = 'hidden'
    
    # Create visible sheets in display order
    build_summary_sheet(wb, timestamp, data_sheet)
    build_sales_list_sheet(wb, timestamp)
    build_items_detail_sheet(wb, timestamp)
    
    # Move data_sheet to the end
    wb.move_sheet("data_sheet", offset=len(wb.sheetnames) - wb.sheetnames.index("data_sheet") - 1)
    
    # Set active sheet
    wb.active = wb["Summary"]
    
    wb.save(report_path)
    print(f"✓ Saved: {report_path}")
    return report_path

if __name__ == "__main__":
  main()

